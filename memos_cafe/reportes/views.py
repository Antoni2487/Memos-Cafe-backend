import logging
import io
from datetime import timedelta
import django
import psutil
from django.contrib.auth import get_user_model
from django.db import connection
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.db.models import Count
from django.db.models import DecimalField
from django.db.models import F
from django.db.models import Sum
from django.db.models import Value
from django.db.models.functions import Coalesce
from django.db.models.functions import TruncDate
from django.utils import timezone
from django.utils.dateparse import parse_date
from rest_framework.exceptions import ValidationError
from rest_framework.response import Response
from rest_framework.views import APIView

from memos_cafe.caja.models import Caja

logger = logging.getLogger("memos_cafe.reportes")
from memos_cafe.caja.models import MovimientoCaja, Pago
from memos_cafe.mesas.models import Mesa
from memos_cafe.ordenes.models import DetalleOrden
from memos_cafe.ordenes.models import Orden
from memos_cafe.utils.permissions import EsAdmin, EsAdminOCajero, modulo_requerido


MAX_DIAS_RANGO_REPORTE = 366


def _parse_fecha(valor, campo):
    """Parsea un query param de fecha (YYYY-MM-DD). Levanta ValidationError
    (400 JSON, capturado automaticamente por APIView.dispatch) en vez de
    dejar que un formato invalido llegue crudo a un filtro __date__gte/lte
    de Django, que revienta con un 500 sin JSON."""
    fecha = parse_date(valor) if valor else None
    if fecha is None:
        raise ValidationError({campo: f"'{valor}' no es una fecha valida. Formato esperado: YYYY-MM-DD."})
    return fecha


def _parse_rango_fechas(request, mensaje_requerido=None):
    """Valida y parsea fecha_inicio/fecha_fin de los query params. Devuelve
    (fecha_inicio, fecha_fin) como date, o levanta ValidationError."""
    fecha_inicio_str = request.query_params.get("fecha_inicio")
    fecha_fin_str = request.query_params.get("fecha_fin")

    if not fecha_inicio_str or not fecha_fin_str:
        raise ValidationError(
            mensaje_requerido or "Los parámetros fecha_inicio y fecha_fin son requeridos."
        )

    fecha_inicio = _parse_fecha(fecha_inicio_str, "fecha_inicio")
    fecha_fin = _parse_fecha(fecha_fin_str, "fecha_fin")
    if fecha_fin < fecha_inicio:
        raise ValidationError({"fecha_fin": "No puede ser anterior a fecha_inicio."})
    if (fecha_fin - fecha_inicio).days > MAX_DIAS_RANGO_REPORTE:
        raise ValidationError(
            {"fecha_fin": f"El rango no puede superar los {MAX_DIAS_RANGO_REPORTE} días."}
        )
    return fecha_inicio, fecha_fin


class DashboardView(APIView):
    """
    GET /api/dashboard/
    KPIs del dia actual para el admin.
    Solo admin.
    """
    permission_classes = [EsAdmin]

    def get(self, request):
        hoy = timezone.localdate()

        # -- Ordenes --------------------------------------------------------
        ordenes_hoy = Orden.objects.filter(fecha_creacion__date=hoy)
        ordenes_abiertas = Orden.objects.filter(estado="abierta").count()
        ordenes_cerradas_hoy = ordenes_hoy.filter(estado="cerrada").count()
        ordenes_anuladas_hoy = ordenes_hoy.filter(estado="anulada").count()

        # -- Ventas del dia ---------------------------------------------------
        pagos_hoy = Pago.objects.filter(
            estado="completado",
            fecha__date=hoy,
        )
        totales = pagos_hoy.aggregate(
            total_ventas=Coalesce(Sum("monto"), Value(0), output_field=DecimalField()),
            total_pagos=Count("id"),
        )
        total_ventas = totales["total_ventas"]
        total_pagos = totales["total_pagos"]
        ticket_promedio = (
            round(total_ventas / total_pagos, 2) if total_pagos > 0 else 0
        )

        # Desglose por metodo de pago (nombre de campo alineado con el
        # frontend: ventas_por_metodo, top-level en la respuesta)
        ventas_por_metodo = list(
            pagos_hoy.values("metodo_pago")
            .annotate(
                total=Coalesce(Sum("monto"), Value(0), output_field=DecimalField()),
                cantidad=Count("id"),
            )
            .order_by("-total")
        )

        # -- Ventas de los ultimos 7 dias (para el grafico de area) ----------
        desde = hoy - timedelta(days=6)
        ventas_por_dia_qs = (
            Pago.objects.filter(estado="completado", fecha__date__gte=desde, fecha__date__lte=hoy)
            .annotate(dia=TruncDate("fecha"))
            .values("dia")
            .annotate(
                total=Coalesce(Sum("monto"), Value(0), output_field=DecimalField()),
                ordenes=Count("orden", distinct=True),
            )
            .order_by("dia")
        )
        por_dia_map = {v["dia"]: v for v in ventas_por_dia_qs}
        ventas_por_dia = []
        for i in range(7):
            dia = desde + timedelta(days=i)
            registro = por_dia_map.get(dia)
            ventas_por_dia.append({
                "fecha": dia.isoformat(),
                "total": registro["total"] if registro else 0,
                "ordenes": registro["ordenes"] if registro else 0,
            })

        # -- Ganancias del mes y del ano --------------------------------------
        inicio_mes = hoy.replace(day=1)
        inicio_anio = hoy.replace(month=1, day=1)

        ventas_mes = Pago.objects.filter(
            estado="completado", fecha__date__gte=inicio_mes, fecha__date__lte=hoy
        ).aggregate(total=Coalesce(Sum("monto"), Value(0), output_field=DecimalField()))["total"]

        ventas_anio = Pago.objects.filter(
            estado="completado", fecha__date__gte=inicio_anio, fecha__date__lte=hoy
        ).aggregate(total=Coalesce(Sum("monto"), Value(0), output_field=DecimalField()))["total"]

        # -- Mesas -------------------------------------------------------------
        mesas = Mesa.objects.filter(activo=True)
        resumen_mesas = {
            "total": mesas.count(),
            "libres": mesas.filter(estado="libre").count(),
            "ocupadas": mesas.filter(estado="ocupada").count(),
            "reservadas": mesas.filter(estado="reservada").count(),
        }

        # -- Caja activa ---------------------------------------------------------
        caja_activa = Caja.objects.get_sesion_abierta()
        caja_info = None
        if caja_activa:
            pagos_caja = caja_activa.pagos.filter(estado="completado")
            ventas_pagos = pagos_caja.aggregate(
                total=Coalesce(Sum("monto"), Value(0), output_field=DecimalField())
            )["total"]
            # Los movimientos manuales (entrada/salida) tambien afectan
            # cuanto hay realmente en caja durante el turno; sin esto
            # 'ventas_turno' no reflejaba salidas como compra de insumos.
            movimientos_neto = MovimientoCaja.objects.neto_por_caja(caja_activa)
            ventas_turno = ventas_pagos + movimientos_neto
            caja_info = {
                "id": caja_activa.id,
                "cajero": caja_activa.usuario.name or caja_activa.usuario.email,
                "fecha_apertura": caja_activa.fecha_apertura,
                "monto_inicial": caja_activa.monto_inicial,
                "ventas_turno": ventas_turno,
            }

        # -- Top 5 productos del dia --------------------------------------------
        top_productos = list(
            DetalleOrden.objects.filter(
                orden__estado="cerrada",
                orden__fecha_cierre__date=hoy,
                producto__isnull=False,
            )
            .values(nombre=F("producto__nombre"))
            .annotate(cantidad=Sum("cantidad"))
            .order_by("-cantidad")[:5]
        )

        return Response({
            "fecha": hoy,
            "ordenes": {
                "abiertas": ordenes_abiertas,
                "cerradas_hoy": ordenes_cerradas_hoy,
                "anuladas_hoy": ordenes_anuladas_hoy,
            },
            "ventas": {
                "total": total_ventas,
                "ticket_promedio": ticket_promedio,
                "por_metodo": ventas_por_metodo,
                "mes": ventas_mes,
                "anio": ventas_anio,
            },
            "ventas_por_dia": ventas_por_dia,
            "ventas_por_metodo": ventas_por_metodo,
            "mesas": resumen_mesas,
            "caja_activa": caja_info,
            "top_productos": top_productos,
        })


class ReporteVentasView(APIView):
    """
    GET /api/reportes/ventas/
    Filtros: fecha_inicio, fecha_fin, usuario_id, metodo_pago
    Solo admin.
    """
    permission_classes = [EsAdmin]

    def get(self, request):
        fecha_inicio, fecha_fin = _parse_rango_fechas(request)
        usuario_id = request.query_params.get("usuario_id")
        metodo_pago = request.query_params.get("metodo_pago")

        # Base queryset de solo pagos completados
        pagos = Pago.objects.filter(
            estado="completado",
            fecha__date__gte=fecha_inicio,
            fecha__date__lte=fecha_fin,
        )

        if usuario_id:
            pagos = pagos.filter(orden__usuario_id=usuario_id)
        if metodo_pago:
            pagos = pagos.filter(metodo_pago=metodo_pago)

        # Totales generales
        totales = pagos.aggregate(
            total_ventas=Coalesce(Sum("monto"), Value(0), output_field=DecimalField()),
            total_ordenes=Count("id"),
        )
        total_ventas = totales["total_ventas"]
        total_ordenes = totales["total_ordenes"]
        ticket_promedio = (
            round(total_ventas / total_ordenes, 2) if total_ordenes > 0 else 0
        )

        # Ventas por día
        ventas_por_dia = list(
            pagos.annotate(fecha_dia=TruncDate("fecha"))
            .values("fecha_dia")
            .annotate(
                total=Coalesce(Sum("monto"), Value(0), output_field=DecimalField()),
                ordenes=Count("id"),
            )
            .order_by("fecha_dia")
        )

        # Ventas por metodo de pago
        ventas_por_metodo = list(
            pagos.values("metodo_pago")
            .annotate(
                total=Coalesce(Sum("monto"), Value(0), output_field=DecimalField()),
                cantidad=Count("id"),
            )
            .order_by("-total")
        )

        logger.info(
            "Reporte ventas consultado | usuario=%s | periodo=%s/%s",
            request.user.email, fecha_inicio, fecha_fin,
        )
        return Response({
            "periodo": {
                "inicio": fecha_inicio,
                "fin": fecha_fin,
            },
            "filtros": {
                "usuario_id": usuario_id,
                "metodo_pago": metodo_pago,
            },
            "total_ventas": total_ventas,
            "total_ordenes": total_ordenes,
            "ticket_promedio": ticket_promedio,
            "ventas_por_dia": ventas_por_dia,
            "ventas_por_metodo_pago": ventas_por_metodo,
        })

    def export_excel(self, request):
        """GET /api/reportes/ventas/export/ — descarga .xlsx"""
        fecha_inicio, fecha_fin = _parse_rango_fechas(request)

        pagos = Pago.objects.filter(
            estado="completado",
            fecha__date__gte=fecha_inicio,
            fecha__date__lte=fecha_fin,
        )

        try:
            wb = openpyxl.Workbook()

            # ── Hoja 1: Resumen ───────────────────────────────────────────────
            ws1 = wb.active
            ws1.title = "Resumen"
            header_font  = Font(bold=True, color="FFFFFF")
            header_fill  = PatternFill("solid", fgColor="2C5545")
            center       = Alignment(horizontal="center")

            ws1.append(["Reporte de Ventas", f"{fecha_inicio} al {fecha_fin}"])
            ws1["A1"].font = Font(bold=True, size=13)
            ws1.append([])

            totales = pagos.aggregate(
                total_ventas=Coalesce(Sum("monto"), Value(0), output_field=DecimalField()),
                total_ordenes=Count("id"),
            )
            tv = totales["total_ventas"]
            to = totales["total_ordenes"]
            ticket = round(tv / to, 2) if to > 0 else 0

            ws1.append(["Total ventas", float(tv)])
            ws1.append(["Total órdenes", to])
            ws1.append(["Ticket promedio", float(ticket)])
            ws1.append([])

            # ── Hoja 2: Ventas por dia ────────────────────────────────────────
            ws2 = wb.create_sheet("Ventas por día")
            headers2 = ["Fecha", "Total (S/)", "Órdenes"]
            ws2.append(headers2)
            for col, h in enumerate(headers2, 1):
                cell = ws2.cell(row=1, column=col)
                cell.font   = header_font
                cell.fill   = header_fill
                cell.alignment = center

            ventas_por_dia = (
                pagos.annotate(fecha_dia=TruncDate("fecha"))
                .values("fecha_dia")
                .annotate(total=Coalesce(Sum("monto"), Value(0), output_field=DecimalField()), ordenes=Count("id"))
                .order_by("fecha_dia")
            )
            for row in ventas_por_dia:
                ws2.append([str(row["fecha_dia"]), float(row["total"]), row["ordenes"]])

            ws2.column_dimensions["A"].width = 14
            ws2.column_dimensions["B"].width = 14
            ws2.column_dimensions["C"].width = 10

            # ── Hoja 3: Por metodo de pago ────────────────────────────────────
            ws3 = wb.create_sheet("Por método de pago")
            headers3 = ["Método", "Total (S/)", "Cantidad"]
            ws3.append(headers3)
            for col, h in enumerate(headers3, 1):
                cell = ws3.cell(row=1, column=col)
                cell.font   = header_font
                cell.fill   = header_fill
                cell.alignment = center

            ventas_por_metodo = (
                pagos.values("metodo_pago")
               .annotate(
                    total=Coalesce(Sum("monto"), Value(0), output_field=DecimalField()),
                    ordenes=Count("id"),
                )
            )
            for row in ventas_por_metodo:
                ws3.append([row["metodo_pago"], float(row["total"]), row["ordenes"]])

            ws3.column_dimensions["A"].width = 16
            ws3.column_dimensions["B"].width = 14
            ws3.column_dimensions["C"].width = 10

            # ── Serializar y responder ────────────────────────────────────────
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            filename = f"reporte_ventas_{fecha_inicio}_{fecha_fin}.xlsx"
            logger.info(
                "Exportación Excel generada | usuario=%s | archivo=%s",
                request.user.email, filename,
            )
            response = HttpResponse(
                buffer.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = f'attachment; filename="{filename}"'
            return response

        except Exception as exc:
            logger.error(
                "Error al generar Excel de ventas | usuario=%s | error=%s",
                request.user.email, str(exc), exc_info=True,
            )
            from rest_framework.response import Response as R
            return R({"detail": "Error al generar el reporte."}, status=500)

class ReporteCajaExportView(APIView):
    """GET /api/reportes/caja/export/"""
    permission_classes = [EsAdmin]

    def get(self, request):
        fecha_inicio, fecha_fin = _parse_rango_fechas(request)

        cajas = Caja.objects.filter(
            estado="cerrada",
            fecha_apertura__date__gte=fecha_inicio,
            fecha_cierre__date__lte=fecha_fin,
        )
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Turnos de Caja"
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="2C5545")
            center      = Alignment(horizontal="center")

            headers = [
                "Caja #", "Cajero", "Estado",
                "Apertura", "Cierre",
                "Monto inicial (S/)", "Total ventas (S/)",
                "Efectivo (S/)", "Tarjeta (S/)", "Yape (S/)", "Plin (S/)",
                "Monto esperado (S/)", "Monto contado (S/)", "Diferencia (S/)",
                "Observaciones",
            ]
            ws.append(headers)
            for col, _ in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center

            for caja in cajas:
                pagos = caja.pagos.filter(estado="completado")
                totales = pagos.aggregate(
                    total=Coalesce(Sum("monto"), Value(0), output_field=DecimalField()),
                )
                por_metodo = {
                    item["metodo_pago"]: float(item["total"])
                    for item in pagos.values("metodo_pago").annotate(
                        total=Coalesce(Sum("monto"), Value(0), output_field=DecimalField())
                    )
                }
                total_ventas         = float(totales["total"])
                monto_inicial        = float(caja.monto_inicial or 0)
                monto_final_esperado = monto_inicial + total_ventas
                monto_final_contado  = float(caja.monto_final or 0)
                diferencia           = (
                    monto_final_contado - monto_final_esperado
                    if caja.monto_final is not None else None
                )
                ws.append([
                    caja.id,
                    caja.usuario.name or caja.usuario.email,
                    caja.estado,
                    str(caja.fecha_apertura) if caja.fecha_apertura else "",
                    str(caja.fecha_cierre)   if caja.fecha_cierre   else "",
                    monto_inicial,
                    total_ventas,
                    por_metodo.get("efectivo", 0),
                    por_metodo.get("tarjeta",  0),
                    por_metodo.get("yape",     0),
                    por_metodo.get("plin",     0),
                    monto_final_esperado,
                    monto_final_contado,
                    diferencia,
                    caja.observaciones or "",
                ])

            for col in ["A","B","C","D","E","F","G","H","I","J","K","L","M","N","O"]:
                ws.column_dimensions[col].width = 18

            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            filename = f"reporte_caja_{fecha_inicio}_{fecha_fin}.xlsx"
            response = HttpResponse(
                buffer.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = f'attachment; filename="{filename}"'
            return response
        except Exception as exc:
            logger.error("Error Excel caja | usuario=%s | error=%s",
                request.user.email, str(exc), exc_info=True)
            return Response({"detail": "Error al generar el reporte."}, status=500)



class ReporteProductosView(APIView):
    """GET /api/reportes/productos/"""
    permission_classes = [EsAdmin]

    def get(self, request):
        fecha_inicio, fecha_fin = _parse_rango_fechas(request)
        usuario_id   = request.query_params.get("usuario_id")
        metodo_pago  = request.query_params.get("metodo_pago")

        detalles = DetalleOrden.objects.filter(
            orden__estado="cerrada",
            orden__fecha_cierre__date__gte=fecha_inicio,
            orden__fecha_cierre__date__lte=fecha_fin,
            producto__isnull=False,
        )
        if usuario_id:
            detalles = detalles.filter(orden__usuario_id=usuario_id)
        if metodo_pago:
            detalles = detalles.filter(orden__pagos__metodo_pago=metodo_pago)

        productos = list(
            detalles.values(
                nombre=F("producto__nombre"),
                categoria=F("producto__categoria__nombre"),
            )
            .annotate(
                cantidad=Sum("cantidad"),
                total=Coalesce(Sum("subtotal"), Value(0), output_field=DecimalField()),
            )
            .order_by("-cantidad")
        )

        detalles_promo = DetalleOrden.objects.filter(
            orden__estado="cerrada",
            orden__fecha_cierre__date__gte=fecha_inicio,
            orden__fecha_cierre__date__lte=fecha_fin,
            promocion__isnull=False,
        )
        if usuario_id:
            detalles_promo = detalles_promo.filter(orden__usuario_id=usuario_id)

        promociones = list(
            detalles_promo.values(nombre=F("promocion__nombre"))
            .annotate(
                cantidad=Sum("cantidad"),
                total=Coalesce(Sum("subtotal"), Value(0), output_field=DecimalField()),
            )
            .order_by("-cantidad")
        )

        return Response({
            "periodo": {"inicio": fecha_inicio, "fin": fecha_fin},
            "productos": productos,
            "promociones": promociones,
        })


class ReporteCajaView(APIView):
    """GET /api/reportes/caja/"""
    permission_classes = [EsAdminOCajero, modulo_requerido("reportes")]

    def get(self, request):
        usuario_id   = request.query_params.get("usuario_id")
        metodo_pago  = request.query_params.get("metodo_pago")
        caja_id      = request.query_params.get("caja_id")
        es_admin     = request.user.groups.filter(name="admin").exists()

        if caja_id:
            try:
                caja = Caja.objects.get(id=caja_id)
            except Caja.DoesNotExist:
                return Response({"detail": "Sesión de caja no encontrada."}, status=404)
            if not es_admin and caja.usuario != request.user:
                return Response({"detail": "No tienes permiso para ver esta sesión."}, status=403)
            return Response(self._cuadre_caja(caja, metodo_pago))

        if not es_admin:
            return Response({"detail": "No tienes permiso para ver este reporte."}, status=403)

        fecha_inicio, fecha_fin = _parse_rango_fechas(
            request, mensaje_requerido="Debe enviar caja_id o fecha_inicio y fecha_fin."
        )

        cajas = Caja.objects.filter(
            estado="cerrada",
            fecha_apertura__date__gte=fecha_inicio,
            fecha_cierre__date__lte=fecha_fin,
        )
        if usuario_id:
            cajas = cajas.filter(usuario_id=usuario_id)

        return Response({
            "periodo": {"inicio": fecha_inicio, "fin": fecha_fin},
            "turnos": [self._cuadre_caja(c, metodo_pago) for c in cajas],
        })

    def _cuadre_caja(self, caja, metodo_pago=None):
        pagos = caja.pagos.filter(estado="completado")
        if metodo_pago:
            pagos = pagos.filter(metodo_pago=metodo_pago)

        totales = pagos.aggregate(
            total=Coalesce(Sum("monto"), Value(0), output_field=DecimalField()),
            cantidad_pagos=Count("id"),
        )
        por_metodo = pagos.values("metodo_pago").annotate(
            total=Coalesce(Sum("monto"), Value(0), output_field=DecimalField()),
            cantidad=Count("id"),
        )
        metodos = {item["metodo_pago"]: item["total"] for item in por_metodo}

        total_ventas         = totales["total"]
        monto_inicial        = caja.monto_inicial or 0
        monto_final_esperado = monto_inicial + total_ventas
        monto_final_contado  = caja.monto_final or 0
        diferencia           = (
            monto_final_contado - monto_final_esperado
            if caja.monto_final is not None else None
        )

        return {
            "caja_id":              caja.id,
            "cajero":               caja.usuario.name or caja.usuario.email,
            "estado":               caja.estado,
            "fecha_apertura":       caja.fecha_apertura,
            "fecha_cierre":         caja.fecha_cierre,
            "monto_inicial":        monto_inicial,
            "total_ventas":         total_ventas,
            "cantidad_pagos":       totales["cantidad_pagos"],
            "desglose_metodos":     metodos,
            "monto_final_esperado": monto_final_esperado,
            "monto_final_contado":  monto_final_contado,
            "diferencia":           diferencia,
            "observaciones":        caja.observaciones,
        }
    
class ReporteProductosExportView(APIView):
    """GET /api/reportes/productos/export/ — descarga Excel de productos y promociones."""
    permission_classes = [EsAdmin]

    def get(self, request):
        fecha_inicio, fecha_fin = _parse_rango_fechas(request)

        productos = list(
            DetalleOrden.objects.filter(
                orden__estado="cerrada",
                orden__fecha_cierre__date__gte=fecha_inicio,
                orden__fecha_cierre__date__lte=fecha_fin,
                producto__isnull=False,
            )
            .values(nombre=F("producto__nombre"), categoria=F("producto__categoria__nombre"))
            .annotate(
                cantidad=Sum("cantidad"),
                total=Coalesce(Sum("subtotal"), Value(0), output_field=DecimalField()),
            )
            .order_by("-cantidad")
        )

        promociones = list(
            DetalleOrden.objects.filter(
                orden__estado="cerrada",
                orden__fecha_cierre__date__gte=fecha_inicio,
                orden__fecha_cierre__date__lte=fecha_fin,
                promocion__isnull=False,
            )
            .values(nombre=F("promocion__nombre"))
            .annotate(
                cantidad=Sum("cantidad"),
                total=Coalesce(Sum("subtotal"), Value(0), output_field=DecimalField()),
            )
            .order_by("-cantidad")
        )

        try:
            wb = openpyxl.Workbook()
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="2C5545")
            center      = Alignment(horizontal="center")

            ws1 = wb.active
            ws1.title = "Productos"
            headers1 = ["Producto", "Categoría", "Cantidad vendida", "Total (S/)"]
            ws1.append(headers1)
            for col, _ in enumerate(headers1, 1):
                cell = ws1.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center
            for p in productos:
                ws1.append([p["nombre"], p["categoria"] or "-", p["cantidad"], float(p["total"])])
            for col in ["A", "B", "C", "D"]:
                ws1.column_dimensions[col].width = 22

            ws2 = wb.create_sheet("Promociones")
            headers2 = ["Promoción", "Cantidad vendida", "Total (S/)"]
            ws2.append(headers2)
            for col, _ in enumerate(headers2, 1):
                cell = ws2.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center
            for p in promociones:
                ws2.append([p["nombre"], p["cantidad"], float(p["total"])])
            for col in ["A", "B", "C"]:
                ws2.column_dimensions[col].width = 22

            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            filename = f"reporte_productos_{fecha_inicio}_{fecha_fin}.xlsx"
            response = HttpResponse(
                buffer.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = f'attachment; filename="{filename}"'
            return response
        except Exception as exc:
            logger.error(
                "Error Excel productos | usuario=%s | error=%s",
                request.user.email, str(exc), exc_info=True,
            )
            return Response({"detail": "Error al generar el reporte."}, status=500)


class ReporteVentasExportView(APIView):
    """GET /api/reportes/ventas/export/ — descarga Excel de ventas."""
    permission_classes = [EsAdmin]

    def get(self, request):
        return ReporteVentasView().export_excel(request)


class ReporteOrdenesView(APIView):
    """
    GET /api/reportes/ordenes/          — datos JSON
    GET /api/reportes/ordenes/export/   — descarga Excel
    Filtros: fecha_inicio, fecha_fin, tipo_orden, estado
    Solo admin.
    """
    permission_classes = [EsAdmin]

    def get(self, request):
        fecha_inicio, fecha_fin = _parse_rango_fechas(request)
        tipo_orden   = request.query_params.get("tipo_orden")
        estado       = request.query_params.get("estado")

        ordenes = Orden.objects.filter(
            fecha_creacion__date__gte=fecha_inicio,
            fecha_creacion__date__lte=fecha_fin,
        ).select_related("mesa", "usuario").prefetch_related("detalles")

        if tipo_orden:
            ordenes = ordenes.filter(tipo_orden=tipo_orden)
        if estado:
            ordenes = ordenes.filter(estado=estado)

        # Resumen por tipo
        resumen_tipo = list(
            ordenes.values("tipo_orden")
            .annotate(
                cantidad=Count("id"),
                total=Coalesce(Sum("total"), Value(0), output_field=DecimalField()),
            )
            .order_by("tipo_orden")
        )

        # Resumen por estado
        resumen_estado = list(
            ordenes.values("estado")
            .annotate(cantidad=Count("id"))
            .order_by("estado")
        )

        # Detalle de ordenes
        detalle = []
        for o in ordenes.order_by("-fecha_creacion"):
            detalle.append({
                "id":           o.id,
                "fecha":        o.fecha_creacion,
                "tipo_orden":   o.tipo_orden,
                "estado":       o.estado,
                "mesa":         o.mesa.numero if o.mesa else None,
                "usuario":      o.usuario.name or o.usuario.email,
                # len(...all()) en vez de .count(): .count() ignora la cache
                # de prefetch_related("detalles") y dispara una query nueva
                # por cada orden listada.
                "items":        len(o.detalles.all()),
                "total":        o.total,
                "cliente":      o.cliente_nombre or None,
                "plataforma":   o.plataforma_delivery or None,
            })

        logger.info(
            "Reporte ordenes consultado | usuario=%s | periodo=%s/%s | total=%d ordenes",
            request.user.email, fecha_inicio, fecha_fin, len(detalle),
        )

        return Response({
            "periodo":        {"inicio": fecha_inicio, "fin": fecha_fin},
            "resumen_tipo":   resumen_tipo,
            "resumen_estado": resumen_estado,
            "ordenes":        detalle,
        })

    def export_excel(self, request):
        """GET /api/reportes/ordenes/export/ — descarga .xlsx"""
        fecha_inicio, fecha_fin = _parse_rango_fechas(request)
        tipo_orden   = request.query_params.get("tipo_orden")
        estado       = request.query_params.get("estado")

        ordenes = Orden.objects.filter(
            fecha_creacion__date__gte=fecha_inicio,
            fecha_creacion__date__lte=fecha_fin,
        ).select_related("mesa", "usuario").prefetch_related("detalles__producto", "detalles__promocion")

        if tipo_orden:
            ordenes = ordenes.filter(tipo_orden=tipo_orden)
        if estado:
            ordenes = ordenes.filter(estado=estado)

        try:
            wb = openpyxl.Workbook()
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="2C5545")
            center      = Alignment(horizontal="center")

            def aplicar_header(ws, headers):
                ws.append(headers)
                for col, _ in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col)
                    cell.font      = header_font
                    cell.fill      = header_fill
                    cell.alignment = center

            # Hoja 1: Resumen
            ws1 = wb.active
            ws1.title = "Resumen"
            ws1.append(["Reporte de Ordenes", f"{fecha_inicio} al {fecha_fin}"])
            ws1["A1"].font = Font(bold=True, size=13)
            ws1.append([])

            resumen_tipo = (
                ordenes.values("tipo_orden")
                .annotate(
                    cantidad=Count("id"),
                    total=Coalesce(Sum("total"), Value(0), output_field=DecimalField()),
                )
                .order_by("tipo_orden")
            )
            ws1.append(["Tipo de orden", "Cantidad", "Total (S/)"])
            for row in resumen_tipo:
                ws1.append([row["tipo_orden"], row["cantidad"], float(row["total"])])

            ws1.append([])
            resumen_estado = (
                ordenes.values("estado")
                .annotate(cantidad=Count("id"))
                .order_by("estado")
            )
            ws1.append(["Estado", "Cantidad"])
            for row in resumen_estado:
                ws1.append([row["estado"], row["cantidad"]])

            ws1.column_dimensions["A"].width = 18
            ws1.column_dimensions["B"].width = 12
            ws1.column_dimensions["C"].width = 14

            # Hoja 2: Detalle de ordenes
            ws2 = wb.create_sheet("Detalle ordenes")
            headers2 = ["#", "Fecha", "Tipo", "Estado", "Mesa", "Usuario", "Cliente", "Plataforma", "Items", "Total (S/)"]
            aplicar_header(ws2, headers2)

            for o in ordenes.order_by("-fecha_creacion"):
                ws2.append([
                    o.id,
                    o.fecha_creacion.strftime("%Y-%m-%d %H:%M"),
                    o.tipo_orden,
                    o.estado,
                    o.mesa.numero if o.mesa else "",
                    o.usuario.name or o.usuario.email,
                    o.cliente_nombre or "",
                    o.plataforma_delivery or "",
                    len(o.detalles.all()),
                    float(o.total),
                ])

            for col in ["A","B","C","D","E","F","G","H","I","J"]:
                ws2.column_dimensions[col].width = 16

            # Hoja 3: Detalle de items
            ws3 = wb.create_sheet("Items por orden")
            headers3 = ["Orden #", "Producto / Promocion", "Cantidad", "Precio unit.", "Subtotal", "Nota", "Impreso"]
            aplicar_header(ws3, headers3)

            for o in ordenes.order_by("id"):
                for d in o.detalles.all():
                    nombre = d.producto.nombre if d.producto else (d.promocion.nombre if d.promocion else "")
                    ws3.append([
                        o.id,
                        nombre,
                        d.cantidad,
                        float(d.precio_unitario),
                        float(d.subtotal),
                        d.nota or "",
                        "Si" if d.impreso else "No",
                    ])

            for col in ["A","B","C","D","E","F","G"]:
                ws3.column_dimensions[col].width = 16

            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            filename = f"reporte_ordenes_{fecha_inicio}_{fecha_fin}.xlsx"
            logger.info(
                "Export Excel ordenes | usuario=%s | archivo=%s",
                request.user.email, filename,
            )
            response = HttpResponse(
                buffer.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = f'attachment; filename="{filename}"' 
            return response

        except Exception as exc:
            logger.error(
                "Error Excel ordenes | usuario=%s | error=%s",
                request.user.email, str(exc), exc_info=True,
            )
            from rest_framework.response import Response as R
            return R({"detail": "Error al generar el reporte."}, status=500)


class ReporteOrdenesExportView(APIView):
    """GET /api/reportes/ordenes/export/"""
    permission_classes = [EsAdmin]

    def get(self, request):
        return ReporteOrdenesView().export_excel(request)


class AlertasView(APIView):
    """GET /api/alertas/ — eventos del turno actual (últimas 8 horas)."""
    permission_classes = [EsAdmin]

    def get(self, request):
        desde = timezone.now() - timedelta(hours=8)

        alertas = []

        # ── Aperturas de caja ─────────────────────────────────────────────
        cajas = Caja.objects.filter(
            fecha_apertura__gte=desde
        ).select_related("usuario")

        for caja in cajas:
            alertas.append({
                "tipo": "caja_apertura",
                "mensaje": f"Caja abierta por el usuario {caja.usuario.name or caja.usuario.email}",
                "fecha": caja.fecha_apertura,
                "icono": "caja",
            })
            if caja.fecha_cierre:
                alertas.append({
                    "tipo": "caja_cierre",
                    "mensaje": f"Caja cerrada por el usuario {caja.usuario.name or caja.usuario.email}",
                    "fecha": caja.fecha_cierre,
                    "icono": "caja",
                })

        # ── Órdenes creadas ───────────────────────────────────────────────
        ordenes = Orden.objects.filter(
            fecha_creacion__gte=desde
        ).select_related("usuario", "mesa")

        for orden in ordenes:
            tipo_display = {
                "mesa":     f"Mesa {orden.mesa.numero}" if orden.mesa else "Mesa",
                "llevar":   "Para llevar",
                "delivery": "Delivery",
            }.get(orden.tipo_orden, orden.tipo_orden)

            alertas.append({
                "tipo": "orden_creada",
                "mensaje": f"Nueva orden #{orden.id} — {tipo_display}",
                "fecha": orden.fecha_creacion,
                "icono": "orden",
            })

        # ── Ventas cobradas ───────────────────────────────────────────────
        pagos = Pago.objects.filter(
            fecha__gte=desde,
            estado="completado",
        ).select_related("orden")

        for pago in pagos:
            alertas.append({
                "tipo": "venta_cobrada",
                "mensaje": f"Venta cobrada — Orden #{pago.orden_id} · S/ {pago.monto}",
                "fecha": pago.fecha,
                "icono": "venta",
            })

        # ── Ordenar por fecha descendente ─────────────────────────────────
        alertas.sort(key=lambda x: x["fecha"], reverse=True)

        # Serializar fechas
        for a in alertas:
            a["fecha"] = a["fecha"].isoformat()

        return Response(alertas)


class HealthCheckView(APIView):
    """
    GET /api/health/
    Health check del sistema — logs, performance y estado general.
    Accesible sin autenticación para monitoreo externo.
    """
    permission_classes = []  # público

    def get(self, request):
        health = {}

        # -- Base de datos --------------------------------------------------
        try:
            connection.ensure_connection()
            health["base_datos"] = {
                "estado": "ok",
                "motor": connection.vendor,
            }
        except Exception as e:
            health["base_datos"] = {
                "estado": "error",
                "detalle": str(e),
            }

        # -- Performance: CPU y memoria --------------------------------------
        health["performance"] = {
            "cpu_porcentaje": psutil.cpu_percent(interval=0.1),
            "memoria_total_mb": round(psutil.virtual_memory().total / 1024 / 1024, 2),
            "memoria_usada_mb": round(psutil.virtual_memory().used / 1024 / 1024, 2),
            "memoria_porcentaje": psutil.virtual_memory().percent,
            "disco_total_gb": round(psutil.disk_usage("/").total / 1024 / 1024 / 1024, 2),
            "disco_usado_gb": round(psutil.disk_usage("/").used / 1024 / 1024 / 1024, 2),
            "disco_porcentaje": psutil.disk_usage("/").percent,
        }

        # -- Health del sistema -----------------------------------------------
        health["sistema"] = {
            "django_version": django.get_version(),
            "estado": "ok",
            "zona_horaria": str(timezone.get_current_timezone()),
        }

        # -- Actividad reciente -------------------------------------------------
        try:
            hoy = timezone.localdate()
            User = get_user_model()
            umbral_activos = timezone.now() - timedelta(minutes=30)
            health["actividad"] = {
                "ordenes_hoy": Orden.objects.filter(
                    fecha_creacion__date=hoy
                ).count(),
                "caja_abierta": Caja.objects.filter(
                    estado="abierta"
                ).exists(),
                "mesas_ocupadas": Mesa.objects.filter(
                    estado="ocupada", activo=True
                ).count(),
                # "Activo" = tuvo un login en los ultimos 30 min (no hay
                # tracking de sesiones vivas al ser JWT sin estado).
                "usuarios_activos_30min": User.objects.filter(
                    last_login__gte=umbral_activos
                ).count(),
            }
        except Exception as e:
            health["actividad"] = {"estado": "error", "detalle": str(e)}

        # -- Estado general -------------------------------------------------
        estado_general = (
            "ok" if health["base_datos"]["estado"] == "ok" else "degradado"
        )

        logger.info(
            "Health check ejecutado | estado=%s | cpu=%s%% | mem=%s%%",
            estado_general,
            health["performance"]["cpu_porcentaje"],
            health["performance"]["memoria_porcentaje"],
        )

        return Response({
            "estado": estado_general,
            "timestamp": timezone.now().isoformat(),
            **health,
        })
